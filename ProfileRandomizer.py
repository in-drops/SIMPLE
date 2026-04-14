

import random
from loguru import logger
import os
from openpyxl import load_workbook
import re


base_dir = os.path.dirname(__file__)
file_path = os.path.join(base_dir, "config", "data", "accounts.xlsx")

wb = load_workbook(file_path)
ws = wb.active

target_header = "Profile Number"
column_index = None

for cell in ws[1]:  # Первая строка — заголовки
    if cell.value == target_header:
        column_index = cell.column
        break

if column_index is None:
    raise ValueError(f"Столбец с заголовком '{target_header}' не найден")

numbers = []
for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
    cell_value = row[0].value
    if cell_value is not None:
        digits_only = re.sub(r"[^\d]", "", str(cell_value))  # удалить всё, кроме цифр
        if digits_only:
            numbers.append(int(digits_only))

def two_pairs():
    random.shuffle(numbers)
    part_size = len(numbers) // 2
    part1 = numbers[:part_size]
    part2 = numbers[part_size:]

    print(f"Часть 1 ({len(part1)}):")
    print(" ".join(map(str, part1)))
    print(f"\nЧасть 2 ({len(part2)}):")
    print(" ".join(map(str, part2)))


def three_pairs():
    random.shuffle(numbers)
    part_size = len(numbers) // 3
    part1 = numbers[:part_size]
    part2 = numbers[part_size:2 * part_size]
    part3 = numbers[2 * part_size:]

    print(f"Часть 1 ({len(part1)}):")
    print(" ".join(map(str, part1)))
    print(f"\nЧасть 2 ({len(part2)}):")
    print(" ".join(map(str, part2)))
    print(f"\nЧасть 3 ({len(part3)}):")
    print(" ".join(map(str, part3)))


def four_pairs():
    random.shuffle(numbers)
    part_size = len(numbers) // 4
    part1 = numbers[:part_size]
    part2 = numbers[part_size:2 * part_size]
    part3 = numbers[2 * part_size:3 * part_size]
    part4 = numbers[3 * part_size:]

    print(f"Часть 1 ({len(part1)}):")
    print(" ".join(map(str, part1)))
    print(f"\nЧасть 2 ({len(part2)}):")
    print(" ".join(map(str, part2)))
    print(f"\nЧасть 3 ({len(part3)}):")
    print(" ".join(map(str, part3)))
    print(f"\nЧасть 4 ({len(part4)}):")
    print(" ".join(map(str, part4)))


def five_pairs():
    random.shuffle(numbers)
    part_size = len(numbers) // 5
    part1 = numbers[:part_size]
    part2 = numbers[part_size:2 * part_size]
    part3 = numbers[2 * part_size:3 * part_size]
    part4 = numbers[3 * part_size:4 * part_size]
    part5 = numbers[4 * part_size:]

    print(f"Часть 1 ({len(part1)}):")
    print(" ".join(map(str, part1)))
    print(f"\nЧасть 2 ({len(part2)}):")
    print(" ".join(map(str, part2)))
    print(f"\nЧасть 3 ({len(part3)}):")
    print(" ".join(map(str, part3)))
    print(f"\nЧасть 4 ({len(part4)}):")
    print(" ".join(map(str, part4)))
    print(f"\nЧасть 5 ({len(part5)}):")
    print(" ".join(map(str, part5)))


def main():
    print(f"\nНайдено профилей в таблице accounts.xlsx: {len(numbers)}\n")

    try:
        parts = int(input('Введите количество равных частей деления профилей (2–5): '))
        print()
        if parts == 2:
            two_pairs()
        elif parts == 3:
            three_pairs()
        elif parts == 4:
            four_pairs()
        elif parts == 5:
            five_pairs()
        else:
            print("Ошибка: можно выбрать только от 2 до 5 частей.")
    except ValueError:
        print("Ошибка: введите целое число от 2 до 5.")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        logger.warning('Программа завершена вручную!')
